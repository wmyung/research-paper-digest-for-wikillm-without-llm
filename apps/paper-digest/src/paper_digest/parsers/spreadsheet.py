from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..models import WorkbookSheet
from ..text import normalize_prose


def _safe_value(value: Any) -> Any:
    if isinstance(value, str):
        return normalize_prose(value)
    return value


def _title(rows: list[list[Any]], sheet_name: str) -> str:
    for row in rows[:25]:
        values = [str(v).strip() for v in row if v not in (None, "")]
        if not values:
            continue
        joined = " ".join(values)
        if len(joined) >= 8:
            return joined[:500]
    return sheet_name


def extract_workbook(path: Path) -> list[WorkbookSheet]:
    workbook = load_workbook(path, read_only=False, data_only=True)
    output: list[WorkbookSheet] = []
    for worksheet in workbook.worksheets:
        rows: list[list[Any]] = []
        nonempty = 0
        for row in worksheet.iter_rows(values_only=True):
            values = [_safe_value(value) for value in row]
            if any(value not in (None, "") for value in values):
                nonempty += sum(value not in (None, "") for value in values)
                while values and values[-1] in (None, ""):
                    values.pop()
                rows.append(values)
        output.append(
            WorkbookSheet(
                file_name=path.name,
                sheet_name=worksheet.title,
                state=worksheet.sheet_state,
                max_row=worksheet.max_row,
                max_column=worksheet.max_column,
                title=_title(rows, worksheet.title),
                rows=rows,
                nonempty_cells=nonempty,
            )
        )
    workbook.close()
    return output
