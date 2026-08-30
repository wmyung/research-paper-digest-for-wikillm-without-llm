from __future__ import annotations

import mimetypes
import re
from hashlib import sha256
from pathlib import Path

import pymupdf
from openpyxl import load_workbook

from .config import DigestConfig
from .models import InputFile
from .parsers.archive import expand_archives

SUPPORTED_SUFFIXES = {
    ".pdf",
    ".xml",
    ".xlsx",
    ".xlsm",
    ".docx",
    ".csv",
    ".tsv",
    ".md",
    ".txt",
    ".rst",
    ".json",
    ".zip",
}


def _hash(path: Path) -> str:
    digest = sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _pdf_pages(path: Path) -> int | None:
    try:
        with pymupdf.open(path) as doc:
            return len(doc)
    except Exception:
        return None


def _sheets(path: Path) -> int | None:
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        value = len(wb.sheetnames)
        wb.close()
        return value
    except Exception:
        return None


def _pdf_score(path: Path) -> float:
    name = path.name.casefold()
    score = 0.0
    if re.search(r"supp|moesm|appendix|supporting|extended|reporting", name):
        score -= 8
    if re.search(r"s\d{4,5}-\d{3}-\d{4,5}-\d", name):
        score += 4
    pages = _pdf_pages(path) or 0
    score += min(pages, 40) / 10
    try:
        with pymupdf.open(path) as doc:
            first = " ".join(page.get_text("text") for page in list(doc)[:2]).casefold()
        if "article" in first or "doi.org/10." in first:
            score += 3
        if "supplementary information" in first or "supporting information" in first:
            score -= 8
        if "received:" in first and "accepted:" in first:
            score += 2
    except Exception:
        score -= 5
    return score


def inventory(paths: list[Path], work_dir: Path, config: DigestConfig) -> tuple[list[InputFile], Path, list[Path]]:
    resolved = [Path(path).resolve() for path in paths]
    for path in resolved:
        if not path.is_file():
            raise ValueError(f"Input file does not exist: {path}")
    unsupported_inputs = [path.name for path in resolved if path.suffix.casefold() not in SUPPORTED_SUFFIXES]
    if unsupported_inputs:
        raise ValueError("Unsupported input type(s): " + ", ".join(unsupported_inputs))
    expanded = expand_archives(resolved, work_dir / "archives", config)
    unsupported_expanded = [
        path.name for path in expanded if path.suffix.casefold() not in SUPPORTED_SUFFIXES - {".zip"}
    ]
    if unsupported_expanded:
        raise ValueError("Unsupported file type(s) inside archive: " + ", ".join(unsupported_expanded))
    pdfs = [path for path in expanded if path.suffix.casefold() == ".pdf"]
    if not pdfs:
        raise ValueError("A canonical full-paper PDF is required.")
    canonical = max(pdfs, key=_pdf_score)

    seen: dict[str, Path] = {}
    files: list[InputFile] = []
    for path in expanded:
        digest = _hash(path)
        duplicate = digest in seen
        role = "duplicate" if duplicate else ("canonical-paper" if path == canonical else "supplement")
        media_type = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
        files.append(
            InputFile(
                path=path,
                role=role,
                media_type=media_type,
                size_bytes=path.stat().st_size,
                sha256=digest,
                page_count=_pdf_pages(path) if path.suffix.casefold() == ".pdf" else None,
                sheet_count=_sheets(path) if path.suffix.casefold() in {".xlsx", ".xlsm"} else None,
                note=f"Duplicate of {seen[digest].name}" if duplicate else None,
            )
        )
        seen.setdefault(digest, path)
    unique = [item.path for item in files if item.role != "duplicate"]
    return files, canonical, unique
